import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

EXCEL_FILE = "jITENDRA.xlsx"
SHEET_NAME = "updated"

# Define available lengths
lengths = [round(x, 1) for x in list(
    [2.0, 2.3, 2.6, 2.9] + [i for i in range(3, 9)]
)]

# Initialize session state
if "page" not in st.session_state:
    st.session_state.page = 1
if "width" not in st.session_state:
    st.session_state.width = None
if "height" not in st.session_state:
    st.session_state.height = None
if "counts" not in st.session_state:
    st.session_state.counts = {l: 0 for l in lengths}

# -------- Screen 1 --------
if st.session_state.page == 1:
    st.title("Step 1: Enter Dimensions")

    width = st.number_input("Enter Width", min_value=0.0, step=0.1)
    height = st.number_input("Enter Height", min_value=0.0, step=0.1)

    if st.button("Next"):
        st.session_state.width = width
        st.session_state.height = height
        st.session_state.page = 2
        st.experimental_rerun()

# -------- Screen 2 --------
elif st.session_state.page == 2:
    st.title("Step 2: Select Lengths")

    # Display buttons for lengths
    for l in lengths:
        col1, col2 = st.columns([1, 3])
        with col1:
            if st.button(f"+ {l}"):
                st.session_state.counts[l] += 1
        with col2:
            st.write(f"Count: {st.session_state.counts[l]}")

    if st.button("Save to Excel"):
        # Prepare dataframe row
        row = {
            "Width": st.session_state.width,
            "Height": st.session_state.height,
        }
        row.update(st.session_state.counts)

        # Append to Excel
        if os.path.exists(EXCEL_FILE):
            book = load_workbook(EXCEL_FILE)
            if SHEET_NAME in book.sheetnames:
                df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
                df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
            else:
                df = pd.DataFrame([row])
        else:
            df = pd.DataFrame([row])

        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

        st.success("Data saved to Excel!")
        st.session_state.page = 1
        st.session_state.counts = {l: 0 for l in lengths}
